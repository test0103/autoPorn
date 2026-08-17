from __future__ import annotations

from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook

from .classifier import Decision

HEADERS = [
    "movie_id",
    "title",
    "tags",
    "action",
    "section_id",
    "section_name",
    "confidence",
    "reason",
    "cover_score",
    "title_score",
    "review_action",
    "review_section",
    "review_title_click",
    "review_cover_click",
    "review_note",
]


def append_decisions(path: str | Path, decisions: Iterable[Decision]) -> None:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.exists():
        wb = load_workbook(path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "operations"
        ws.append(HEADERS)
    for decision in decisions:
        ws.append([
            decision.movie.id,
            decision.movie.name,
            ",".join(decision.movie.tag_names),
            decision.action,
            decision.section.id if decision.section else "",
            decision.section.name if decision.section else "",
            decision.confidence,
            decision.reason,
            decision.cover_score,
            decision.title_score,
            "",
            "",
            "",
            "",
            "",
        ])
    wb.save(path)


def read_review_training(path: str | Path) -> list[tuple[str, str]]:
    path = Path(path)
    if not path.exists():
        return []
    wb = load_workbook(path)
    ws = wb.active
    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {name: i for i, name in enumerate(headers)}
    rows: list[tuple[str, str]] = []
    def cell(row: tuple, name: str) -> str:
        column = idx.get(name)
        if column is None or column >= len(row):
            return ""
        return str(row[column] or "")

    for row in ws.iter_rows(min_row=2, values_only=True):
        review_action = cell(row, "review_action") or cell(row, "approved")
        approved = review_action.lower() in {"1", "true", "yes", "y", "通过", "approve", "keep"}
        text = cell(row, "title") + " " + cell(row, "tags")
        section = cell(row, "review_section") or cell(row, "correct_section") or cell(row, "section_name")
        title_click = cell(row, "review_title_click")
        cover_click = cell(row, "review_cover_click")
        if approved and text.strip() and section.strip():
            rows.append((" ".join([text, title_click, cover_click]).strip(), section))
    return rows
